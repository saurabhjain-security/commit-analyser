#!/usr/bin/env python3
"""
Commit Analyser
---------------
Script  : Commit Analyser
Author  : Saurabh Jain
Version : 4.0

What changed from v3.0 → v4.0  (all 31 issues from the audit fixed)
  BUG FIXES
    BUG-01  DATE_FILE moved inside main() — no longer evaluated at import time
    BUG-02  _get() now has MAX_RETRIES=5 cap — no more infinite retry loops
    BUG-03  403 handling split: only sleeps on true rate-limit (RL-remaining=0);
            permission 403 fails fast with a clear message
    BUG-04  b180 / summary bucket label now honours --stale-days correctly
    BUG-05  Personal repo dedup implemented via full_name set (no double-counting)
    BUG-06  --since documented as "filter API results"; Days Inactive always
            measures from today so the column is meaningful
    BUG-07  Rate-limit reset shown in UTC via datetime.fromtimestamp(..., tz=UTC)
    BUG-08  Each thread uses its own requests.Session — no shared state race
    BUG-09  --filter active excludes archived repos, not repos with stale="No"

  SILENT DATA FIXES
    LOSS-01 Excel row colouring applies only to the Stale column cell, not the URL
    LOSS-02 JSON keys use consistent snake_case mapping from a fixed dict
    LOSS-03 Summary column header is dynamic: "31–{stale_days}d"
    LOSS-04 Archived repos skip the commits API call entirely (saves quota)

  PERFORMANCE
    PERF-01 Repo listing is concurrent — all orgs fetched in parallel threads
    PERF-02 Workers auto-throttle when RL-remaining drops below 200
    PERF-03 Excel written in one pass (openpyxl only, no pandas intermediate)

  NEW FEATURES
    • --verbose / -v     Step-level logging: API URLs, status codes, timings
    • --exclude-archived Skip archived repos from scan entirely
    • Ctrl+C handling    Graceful shutdown; partial results saved before exit
    • --sort             Sort output by: days (default) | name | org | stale
    • Per-org Excel sheets (one sheet per org + "All Repos" summary sheet)
    • Token scope check  Warns if token lacks repo or read:org at startup

  CODE QUALITY
    • NO_COMMIT = 99_999 named constant replaces magic number throughout
    • _c() colour helper is module-level (not redefined in a loop)
    • RepoRow TypedDict documents the row schema explicitly
    • print_summary() receives a string, not the full GHClient
    • Network/HTTP errors are raised by _get(); callers decide display

Output Columns (terminal + Excel):
  Sno | Org / Owner | Repo Name | Repo URL | Fork | Archived | Last Commit (days) | Stale

  "Last Commit (days)" column contains only the integer number of days (e.g. 818),
  NOT "818 days ago".  Terminal uses coloured "818 d" label; exports use the plain integer.

Pre-requisite:
  pip install rich requests pandas openpyxl

  Token (one of):
    export GITHUB_TOKEN=ghp_...
    export GH_TOKEN=ghp_...
    gh auth login   (fallback)

Usage:
  python commit_analyser.py
  python commit_analyser.py --org acme-corp another-org
  python commit_analyser.py --stale-days 90
  python commit_analyser.py --exclude-forks --exclude-archived
  python commit_analyser.py --filter stale --output-format csv
  python commit_analyser.py --include-personal --no-export
  python commit_analyser.py --since 2024-01-01
  python commit_analyser.py --sort name
  python commit_analyser.py --verbose
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import signal
import subprocess
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed, Future
from datetime import datetime, timezone
from threading import Lock, local as thread_local
from typing import Dict, List, Optional, Set, Tuple, TypedDict

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rich.console import Console
from rich.panel import Panel
from rich.progress import (
    BarColumn,
    MofNCompleteColumn,
    Progress,
    SpinnerColumn,
    TaskProgressColumn,
    TextColumn,
    TimeElapsedColumn,
)
from rich.rule import Rule
from rich.table import Table
from rich.traceback import install as install_rich_traceback
from rich import box

install_rich_traceback(show_locals=False)

# ── Constants ─────────────────────────────────────────────────────────────────
GITHUB_API    = "https://api.github.com"
VERSION       = "4.0"
AUTHOR        = "Saurabh Jain"
MAX_RETRIES   = 5          # BUG-02: cap on _get() retry attempts
NO_COMMIT     = 99_999     # QUAL-03: named sentinel for "no commits found"
RL_THROTTLE   = 200        # PERF-02: slow down when remaining calls drop below this

console    = Console()
print_lock = Lock()        # serialise Rich output from worker threads
_shutdown  = False         # Ctrl+C flag


# ── TypedDict schema ──────────────────────────────────────────────────────────

class RepoRow(TypedDict):
    """QUAL-06: explicit schema replaces opaque dict."""
    org_name:    str
    repo_name:   str
    repo_url:    str
    is_fork:     bool
    is_archived: bool
    last_commit_days: int        # integer days; NO_COMMIT when no commits exist
    last_commit_iso:  str        # raw ISO date string, empty when none
    stale:       str             # "Yes" | "No" | "N/A" | "No commits"
    _sort_key:   int             # same as last_commit_days, used internally


# ── Excel styling (module-level, immutable) ───────────────────────────────────
_HDR_FILL      = PatternFill("solid", start_color="1F3864")
_HDR_FONT      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_BODY_FONT     = Font(name="Calibri", size=10)
_THIN          = Side(style="thin", color="D9D9D9")
_BORDER        = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)
_FILL_STALE    = PatternFill("solid", start_color="FFC7CE")
_FILL_FRESH    = PatternFill("solid", start_color="C6EFCE")
_FILL_WARN     = PatternFill("solid", start_color="FFEB9C")
_FILL_ARCHIVED = PatternFill("solid", start_color="D9D9D9")
_FILL_NONE     = PatternFill("solid", start_color="FFFFFF")

VERBOSE = False   # set from --verbose flag in main()


# ── Logging helpers ───────────────────────────────────────────────────────────

def vlog(msg: str) -> None:
    """Print a verbose diagnostic line (only when --verbose is active)."""
    if not VERBOSE:
        return
    ts = datetime.now(timezone.utc).strftime("%H:%M:%S")
    with print_lock:
        console.print(f"  [dim cyan]VRB {ts}[/]  {msg}")


# ── Banner ────────────────────────────────────────────────────────────────────

def print_banner(args: argparse.Namespace) -> None:
    art = (
        "  ██████╗ ██████╗ ███╗   ███╗███╗   ███╗██╗████████╗\n"
        " ██╔════╝██╔═══██╗████╗ ████║████╗ ████║██║╚══██╔══╝\n"
        " ██║     ██║   ██║██╔████╔██║██╔████╔██║██║   ██║   \n"
        " ██║     ██║   ██║██║╚██╔╝██║██║╚██╔╝██║██║   ██║   \n"
        " ╚██████╗╚██████╔╝██║ ╚═╝ ██║██║ ╚═╝ ██║██║   ██║   \n"
        "  ╚═════╝ ╚═════╝ ╚═╝     ╚═╝╚═╝     ╚═╝╚═╝   ╚═╝   \n"
        "\n"
        " █████╗ ███╗   ██╗ █████╗ ██╗  ██╗   ██╗███████╗███████╗██████╗ \n"
        "██╔══██╗████╗  ██║██╔══██╗██║  ╚██╗ ██╔╝██╔════╝██╔════╝██╔══██╗\n"
        "███████║██╔██╗ ██║███████║██║   ╚████╔╝ ███████╗█████╗  ██████╔╝\n"
        "██╔══██║██║╚██╗██║██╔══██║██║    ╚██╔╝  ╚════██║██╔══╝  ██╔══██╗\n"
        "██║  ██║██║ ╚████║██║  ██║███████╗██║   ███████║███████╗██║  ██║\n"
        "╚═╝  ╚═╝╚═╝  ╚═══╝╚═╝  ╚═╝╚══════╝╚═╝   ╚══════╝╚══════╝╚═╝  ╚═╝"
    )
    flags = []
    if args.exclude_forks:    flags.append("forks excluded")
    if args.exclude_archived: flags.append("archived excluded")
    if args.include_personal: flags.append("personal included")
    if args.since:            flags.append(f"since {args.since}")
    if VERBOSE:               flags.append("[bold cyan]VERBOSE[/]")
    flags_str = "  [dim]·[/]  ".join(flags) if flags else ""

    console.print()
    console.print(Panel(
        f"[bold bright_cyan]{art}[/]\n\n"
        f"  [bold bright_green]GitHub Organisation Repository Commit Analyser[/]"
        f"   [dim]|[/]   [dim]v{VERSION}[/]"
        f"   [dim]|[/]   [bold yellow]Author: {AUTHOR}[/]\n\n"
        f"  [dim cyan]Stale threshold :[/] [bold white]{args.stale_days} days[/]\n"
        f"  [dim cyan]Workers         :[/] [bold white]{args.workers}[/]"
        + (f"\n  [dim cyan]Flags           :[/] {flags_str}" if flags_str else ""),
        border_style="bright_blue",
        box=box.DOUBLE_EDGE,
        expand=False,
        padding=(1, 2),
    ))
    console.print()


# ── CLI ───────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        prog="commit_analyser",
        description=f"GitHub Repository Commit Analyser v{VERSION}",
        epilog=(
            "Examples:\n"
            "  python commit_analyser.py\n"
            "  python commit_analyser.py --org acme-corp\n"
            "  python commit_analyser.py --stale-days 90 --exclude-forks\n"
            "  python commit_analyser.py --filter stale --output-format csv\n"
            "  python commit_analyser.py --include-personal --no-export\n"
            "  python commit_analyser.py --since 2024-01-01\n"
            "  python commit_analyser.py --sort name\n"
            "  python commit_analyser.py --verbose\n"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--org", nargs="+", metavar="ORG", default=None,
                   help="Target specific org(s). Default: all orgs the token belongs to.")
    p.add_argument("--stale-days", dest="stale_days", type=int, default=180, metavar="DAYS",
                   help="Days inactive before a repo is considered stale (default: 180).")
    p.add_argument("--output", "-o", default="",
                   help="Output file path (default: commit_analyser_<timestamp>.<ext>).")
    p.add_argument("--output-format", dest="output_format",
                   choices=["xlsx", "csv", "json"], default="xlsx",
                   help="Export format: xlsx (default) | csv | json.")
    p.add_argument("--no-export", dest="no_export", action="store_true",
                   help="Skip file export — terminal output only.")
    p.add_argument("--include-personal", dest="include_personal", action="store_true",
                   help="Also scan repos in the authenticated user's personal namespace.")
    p.add_argument("--exclude-forks", dest="exclude_forks", action="store_true",
                   help="Skip forked repositories.")
    p.add_argument("--exclude-archived", dest="exclude_archived", action="store_true",
                   help="Skip archived repositories entirely (saves API quota).")
    p.add_argument("--filter", dest="filter_mode",
                   choices=["all", "stale", "archived", "active"], default="all",
                   help="Display subset: all (default) | stale | archived | active.")
    p.add_argument("--since", default=None, metavar="YYYY-MM-DD",
                   help="Only fetch commits after this date.")
    p.add_argument("--workers", type=int, default=10, metavar="N",
                   help="Parallel workers for fetching commit dates (default: 10).")
    p.add_argument("--sort", dest="sort_by",
                   choices=["days", "name", "org", "stale"], default="days",
                   help="Sort results by: days (default) | name | org | stale.")
    p.add_argument("--verbose", "-v", action="store_true", default=False,
                   help="Show per-call API logging and step timings.")
    return p.parse_args()


# ── Auth ──────────────────────────────────────────────────────────────────────

def get_token() -> str:
    for env_var in ("GITHUB_TOKEN", "GH_TOKEN"):
        token = os.environ.get(env_var, "").strip()
        if token:
            console.print(f"  [green]✓[/]  Token from env var [bold cyan]{env_var}[/]")
            return token
    try:
        r = subprocess.run(["gh", "auth", "token"],
                           capture_output=True, text=True, timeout=10)
    except FileNotFoundError:
        console.print(Panel(
            "[bold red]✗  No token found[/]\n\n"
            "  Set [bold cyan]GITHUB_TOKEN[/] or [bold cyan]GH_TOKEN[/] env var,\n"
            "  or install gh CLI and run [bold cyan]gh auth login[/]",
            border_style="red", box=box.ROUNDED, expand=False))
        sys.exit(1)
    except subprocess.TimeoutExpired:
        console.print("  [bold red]✗[/]  `gh auth token` timed out.")
        sys.exit(1)

    token = r.stdout.strip()
    if not token or r.returncode != 0:
        console.print(Panel(
            "[bold red]✗  Not authenticated[/]\n\n"
            "  Set [bold cyan]GITHUB_TOKEN[/] env var, or run [bold cyan]gh auth login[/]",
            border_style="red", box=box.ROUNDED, expand=False))
        sys.exit(1)
    console.print("  [green]✓[/]  Token via [bold cyan]gh auth token[/]")
    return token


# ── GitHub API Client ─────────────────────────────────────────────────────────

# BUG-08: Thread-local storage so each worker thread gets its own Session
_tl = thread_local()


def _make_session(token: str) -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "Authorization":        f"Bearer {token}",
        "Accept":               "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    })
    return s


class GHClient:
    """
    Thread-safe GitHub REST API client.
    Each worker thread gets its own requests.Session via thread-local storage.
    Rate-limit state is tracked in shared (lock-protected) attributes.
    """

    def __init__(self, token: str) -> None:
        self._token            = token
        self._rl_remaining     = 5000
        self._rl_limit         = 5000
        self._rl_lock          = Lock()

    # ── Per-thread session ─────────────────────────────────────────────
    def _session(self) -> requests.Session:
        if not hasattr(_tl, "session"):
            _tl.session = _make_session(self._token)
        return _tl.session

    # ── Core GET (BUG-02: retry cap; BUG-03: 403 split; BUG-07: UTC) ──
    def _get(
        self,
        path: str,
        params: Optional[Dict] = None,
    ) -> requests.Response:
        url     = f"{GITHUB_API}{path}" if path.startswith("/") else path
        backoff = 2
        attempt = 0

        while attempt < MAX_RETRIES:
            # PERF-02: auto-throttle when RL running low
            with self._rl_lock:
                remaining = self._rl_remaining
            if remaining < RL_THROTTLE and remaining > 0:
                vlog(f"RL low ({remaining}) — pausing 2s before {url[:60]}")
                time.sleep(2)

            try:
                t0   = time.monotonic()
                resp = self._session().get(url, params=params, timeout=30)
                ms   = (time.monotonic() - t0) * 1000
            except (requests.ConnectionError, requests.Timeout) as exc:
                attempt += 1
                vlog(f"Network error ({exc.__class__.__name__}) attempt {attempt}/{MAX_RETRIES} — "
                     f"retry in {backoff}s")
                time.sleep(backoff)
                backoff = min(backoff * 2, 60)
                continue

            # Update shared rate-limit counters
            try:
                with self._rl_lock:
                    self._rl_remaining = int(
                        resp.headers.get("X-RateLimit-Remaining", self._rl_remaining))
                    self._rl_limit = int(
                        resp.headers.get("X-RateLimit-Limit", self._rl_limit))
            except (ValueError, TypeError):
                pass

            vlog(
                f"[dim]GET[/] {url[:80]}  "
                f"[{'bold green' if resp.ok else 'bold red'}]{resp.status_code}[/]  "
                f"[dim]{ms:.0f}ms  RL={self._rl_remaining}[/]"
            )

            # BUG-03: true rate-limit only when RL-remaining hits zero
            is_rate_limited = (
                resp.status_code == 429
                or (resp.status_code == 403
                    and resp.headers.get("X-RateLimit-Remaining") == "0")
            )
            if is_rate_limited:
                reset = int(resp.headers.get("X-RateLimit-Reset", time.time() + 60))
                wait  = max(reset - int(time.time()), 1)
                # BUG-07: always show UTC
                reset_utc = datetime.fromtimestamp(reset, tz=timezone.utc).strftime("%H:%M:%S UTC")
                with print_lock:
                    console.print(
                        f"  [bold yellow]⚠[/]  Rate limit — "
                        f"waiting [bold]{wait}s[/]  [dim](resets {reset_utc})[/]"
                    )
                time.sleep(wait + 1)
                continue   # don't count toward attempt limit

            # 403 that is NOT a rate-limit = permission error → fail immediately
            if resp.status_code == 403:
                try:
                    msg = resp.json().get("message", resp.text[:120])
                except Exception:
                    msg = resp.text[:120]
                vlog(f"[bold red]403 permission denied[/] {url[:80]}  {msg}")
                return resp   # caller sees resp.ok == False

            # 5xx server error: back off and retry
            if resp.status_code >= 500:
                attempt += 1
                vlog(f"Server error {resp.status_code} — retry {attempt}/{MAX_RETRIES} in {backoff}s")
                time.sleep(backoff)
                backoff = min(backoff * 2, 60)
                continue

            return resp

        # Exhausted retries
        raise RuntimeError(f"Gave up after {MAX_RETRIES} attempts: {url}")

    # ── Pagination ─────────────────────────────────────────────────────
    def paginate(self, path: str, params: Optional[Dict] = None) -> Tuple[List[dict], bool]:
        """Returns (items, complete).  complete=False = a page errored."""
        p = dict(params or {})
        p.setdefault("per_page", 100)
        p["page"] = 1
        collected: List[dict] = []
        while True:
            try:
                resp = self._get(path, params=p)
            except RuntimeError as exc:
                with print_lock:
                    console.print(f"  [bold red]✗[/]  {exc}")
                return collected, False
            if not resp.ok:
                with print_lock:
                    console.print(
                        f"  [bold red]✗[/]  API {resp.status_code} on {path} — partial results")
                return collected, False
            items = resp.json()
            if not isinstance(items, list) or not items:
                return collected, True
            collected.extend(items)
            if 'rel="next"' not in resp.headers.get("Link", ""):
                return collected, True
            p["page"] += 1

    # ── Lookups ────────────────────────────────────────────────────────
    def get_one(self, path: str) -> Optional[dict]:
        try:
            resp = self._get(path)
            return resp.json() if resp.ok else None
        except RuntimeError:
            return None

    def rate_summary(self) -> str:
        with self._rl_lock:
            return f"{self._rl_remaining:,}/{self._rl_limit:,}"

    # ── Rate-limit pre-flight ──────────────────────────────────────────
    def check_rate_limit(self) -> None:
        data = self.get_one("/rate_limit")
        if not data:
            return
        rate      = data.get("rate", {})
        remaining = rate.get("remaining", 0)
        limit     = rate.get("limit", 5000)
        reset_ts  = rate.get("reset", 0)
        reset_utc = datetime.fromtimestamp(reset_ts, tz=timezone.utc).strftime("%H:%M:%S UTC")
        with self._rl_lock:
            self._rl_remaining = remaining
            self._rl_limit     = limit
        colour = "green" if remaining > 1000 else ("yellow" if remaining > 200 else "red")
        console.print(
            f"  [green]✓[/]  Rate limit: "
            f"[bold {colour}]{remaining:,}[/] / {limit:,} calls remaining  "
            f"[dim](resets {reset_utc})[/]"
        )
        if remaining < 100:
            console.print(
                "  [bold red]⚠  Fewer than 100 API calls remaining.[/]  "
                "Scan may be incomplete — wait for reset or use a different token."
            )

    # ── Token scope check ──────────────────────────────────────────────
    def check_scopes(self) -> None:
        """FEAT-07: warn if token lacks repo or read:org scope."""
        try:
            resp = self._get("/user")
            scopes = resp.headers.get("X-OAuth-Scopes", "")
            if scopes == "" and "Bearer" in self._session().headers.get("Authorization", ""):
                return   # fine-grained token — no scope header
            needed = {"repo", "read:org"}
            have   = {s.strip() for s in scopes.split(",")}
            if not needed.issubset(have):
                missing = needed - have
                console.print(
                    f"  [bold yellow]⚠[/]  Token may be missing scope(s): "
                    f"[bold]{', '.join(missing)}[/]  "
                    f"[dim](private repos or org listing may fail silently)[/]"
                )
        except Exception:
            pass


# ── Helpers ───────────────────────────────────────────────────────────────────

def days_since(iso_date: Optional[str]) -> Tuple[int, str]:
    """
    Returns (days_int, iso_str).
    days_int = NO_COMMIT when no commits exist or date is unparseable.
    iso_str  = the raw ISO string, or "" when unavailable.
    """
    if not iso_date:
        return NO_COMMIT, ""
    try:
        dt   = datetime.fromisoformat(iso_date.replace("Z", "+00:00"))
        now  = datetime.now(timezone.utc)
        days = (now - dt).days
        return days, iso_date
    except Exception:
        return NO_COMMIT, ""


def days_display(days: int) -> str:
    """
    Human-readable terminal label — integer only, no "days ago" suffix.
    Examples: "Today", "1 d", "818 d", "No commits"
    """
    if days == NO_COMMIT:
        return "No commits"
    if days == 0:
        return "Today"
    return f"{days} d"


def commit_colour(days: int) -> str:
    """Rich markup colour for a days value."""
    label = days_display(days)
    if days == NO_COMMIT:
        return f"[dim white]{label}[/]"
    if days == 0:
        return f"[bold green]{label}[/]"
    if days <= 7:
        return f"[bold green]{label}[/]"
    if days <= 30:
        return f"[bold yellow]{label}[/]"
    if days <= 180:
        return f"[bold orange3]{label}[/]"
    return f"[bold red]{label}[/]"


def _c(val: int, color: str) -> str:
    """QUAL-01: module-level colour helper (not redefined per loop iteration)."""
    return f"[{color}]{val}[/]" if val else "[dim]0[/]"


# ── Data collection ───────────────────────────────────────────────────────────

def get_last_commit_date(
    client: GHClient,
    full_name: str,
    default_branch: str,
    since: Optional[str],
) -> Optional[str]:
    """Returns commit.author.date ISO string, or None."""
    params: Dict = {"per_page": 1}
    if default_branch:
        params["sha"] = default_branch
    if since:
        params["since"] = since

    try:
        resp = client._get(f"/repos/{full_name}/commits", params=params)
    except RuntimeError:
        return None

    if not resp.ok:
        # Retry without sha param if branch not found
        if "sha" in params:
            params2 = {k: v for k, v in params.items() if k != "sha"}
            try:
                resp = client._get(f"/repos/{full_name}/commits", params=params2)
            except RuntimeError:
                return None
            if not resp.ok:
                return None
        else:
            return None

    data = resp.json()
    if isinstance(data, list) and data:
        return (
            data[0]
            .get("commit", {})
            .get("author", {})
            .get("date")
        )
    return None


def fetch_repo_row(
    client: GHClient,
    org: str,
    repo: dict,
    stale_days: int,
    since: Optional[str],
) -> RepoRow:
    """Build one RepoRow for a single repo (called from thread pool)."""
    if _shutdown:
        raise InterruptedError("Shutdown requested")

    repo_name      = repo.get("name", "")
    full_name      = repo.get("full_name", "")
    default_branch = repo.get("default_branch") or ""
    html_url       = repo.get("html_url") or f"https://github.com/{full_name}"
    is_archived    = bool(repo.get("archived", False))
    is_fork        = bool(repo.get("fork", False))

    # LOSS-04: skip commits API call entirely for archived repos
    if is_archived:
        return RepoRow(
            org_name=org, repo_name=repo_name, repo_url=html_url,
            is_fork=is_fork, is_archived=True,
            last_commit_days=NO_COMMIT, last_commit_iso="",
            stale="N/A", _sort_key=NO_COMMIT,
        )

    iso_date = get_last_commit_date(client, full_name, default_branch, since)
    days, iso_str = days_since(iso_date)

    if days == NO_COMMIT:
        stale_val = "No commits"
    elif days > stale_days:
        stale_val = "Yes"
    else:
        stale_val = "No"

    return RepoRow(
        org_name=org, repo_name=repo_name, repo_url=html_url,
        is_fork=is_fork, is_archived=False,
        last_commit_days=days, last_commit_iso=iso_str,
        stale=stale_val, _sort_key=days,
    )


def fetch_org_repos(
    client: GHClient,
    org: str,
    exclude_forks: bool,
    exclude_archived: bool,
) -> List[dict]:
    """Fetch all repos for one org with optional filters."""
    repos, _ = client.paginate(f"/orgs/{org}/repos", params={"type": "all"})
    out = []
    for r in repos:
        if exclude_forks and r.get("fork"):
            continue
        if exclude_archived and r.get("archived"):
            continue
        out.append(r)
    return out


# ── Sorting ───────────────────────────────────────────────────────────────────

def sort_rows(rows: List[RepoRow], sort_by: str) -> List[RepoRow]:
    if sort_by == "name":
        return sorted(rows, key=lambda r: r["repo_name"].lower())
    if sort_by == "org":
        return sorted(rows, key=lambda r: (r["org_name"].lower(), r["repo_name"].lower()))
    if sort_by == "stale":
        order = {"Yes": 0, "No commits": 1, "No": 2, "N/A": 3}
        return sorted(rows, key=lambda r: (order.get(r["stale"], 9), r["_sort_key"]))
    # default: days (most recent first, no-commits sink to bottom)
    return sorted(rows, key=lambda r: r["_sort_key"])


# ── Filter ────────────────────────────────────────────────────────────────────

def apply_filter(rows: List[RepoRow], mode: str) -> List[RepoRow]:
    # BUG-09: "active" = not stale and not archived
    if mode == "stale":
        return [r for r in rows if r["stale"] == "Yes"]
    if mode == "archived":
        return [r for r in rows if r["is_archived"]]
    if mode == "active":
        return [r for r in rows if not r["is_archived"] and r["stale"] not in ("Yes", "No commits")]
    return rows


# ── Terminal table ────────────────────────────────────────────────────────────

def print_results_table(rows: List[RepoRow], filter_mode: str, stale_days: int) -> None:
    display = apply_filter(rows, filter_mode)

    console.print(Rule("[bold bright_cyan]  Scan Results  [/]", style="bright_blue"))
    console.print()

    filter_label = "" if filter_mode == "all" else f" — filter: [bold]{filter_mode}[/]"
    tbl = Table(
        title=(
            f"[bold bright_cyan]GitHub Repository Commit Analysis[/]  "
            f"[dim]({len(display)} repos{filter_label})[/]"
        ),
        box=box.ROUNDED,
        border_style="bright_blue",
        header_style="bold bright_blue",
        show_lines=True,
        expand=True,
    )
    tbl.add_column("Sno",              justify="right",  style="dim white",  width=5,    no_wrap=True)
    tbl.add_column("Org / Owner",      justify="left",   style="bold cyan",  min_width=16, no_wrap=True)
    tbl.add_column("Repo Name",        justify="left",   style="white",      min_width=18)
    tbl.add_column("Repo URL",         justify="left",   overflow="fold",    min_width=34)
    tbl.add_column("Fork",             justify="center", style="white",      width=7,    no_wrap=True)
    tbl.add_column("Archived",         justify="center", style="white",      width=10,   no_wrap=True)
    tbl.add_column("Last Commit (d)",  justify="center", style="white",      width=15,   no_wrap=True)
    tbl.add_column("Stale",            justify="center", style="white",      width=10,   no_wrap=True)

    for idx, row in enumerate(display, start=1):
        is_stale    = row["stale"] == "Yes"
        url_markup  = (
            f"[bold red]{row['repo_url']}[/]" if is_stale
            else f"[cyan]{row['repo_url']}[/]"
        )
        archived_markup = "[bold yellow]True[/]"  if row["is_archived"] else "[dim white]False[/]"
        fork_markup     = "[dim white]True[/]"    if row["is_fork"]     else "[dim white]False[/]"
        commit_markup   = commit_colour(row["last_commit_days"])

        stale_val = row["stale"]
        if stale_val == "Yes":         stale_markup = "[bold red]Yes[/]"
        elif stale_val == "N/A":       stale_markup = "[dim white]N/A[/]"
        elif stale_val == "No commits":stale_markup = "[dim white]—[/]"
        else:                          stale_markup = "[green]No[/]"

        tbl.add_row(
            str(idx),
            row["org_name"],
            row["repo_name"],
            url_markup,
            fork_markup,
            archived_markup,
            commit_markup,
            stale_markup,
        )

    console.print(tbl)
    console.print()


# ── Summary table ─────────────────────────────────────────────────────────────

def print_summary(
    rows: List[RepoRow],
    org_labels: List[str],
    elapsed: float,
    stale_days: int,
    rate_summary: str,   # QUAL-04: pass string, not GHClient
) -> None:
    console.print(Rule("[bold bright_green]  Summary  [/]", style="bright_green"))
    console.print()

    # BUG-04 / LOSS-03: dynamic column label honours stale_days
    b_mid_label = f"31–{stale_days}d"

    per_org = Table(
        title="[bold white]Per-Organisation Breakdown[/]",
        box=box.ROUNDED, border_style="bright_green",
        header_style="bold bright_green", show_lines=True, expand=False,
    )
    per_org.add_column("Organisation",       style="bold cyan", no_wrap=True, min_width=20)
    per_org.add_column("Scanned",            justify="right",   style="white", width=9)
    per_org.add_column("≤7d",                justify="right",   width=7)
    per_org.add_column("8–30d",              justify="right",   width=7)
    per_org.add_column(b_mid_label,          justify="right",   width=10)
    per_org.add_column(f">{stale_days}d",    justify="right",   width=9)
    per_org.add_column("No Commits",         justify="right",   width=11)
    per_org.add_column("Archived",           justify="right",   width=9)
    per_org.add_column("Forks",              justify="right",   width=7)

    grand: Dict[str, int] = dict(
        scanned=0, b7=0, b30=0, b_mid=0, stale=0, no_cmt=0, archived=0, forks=0
    )

    for org in org_labels:
        org_rows = [r for r in rows if r["org_name"] == org]
        if not org_rows:
            continue
        scanned  = len(org_rows)
        b7       = sum(1 for r in org_rows if 0 <= r["_sort_key"] <= 7)
        b30      = sum(1 for r in org_rows if 8 <= r["_sort_key"] <= 30)
        b_mid    = sum(1 for r in org_rows if 31 <= r["_sort_key"] <= stale_days)
        stale    = sum(1 for r in org_rows if r["stale"] == "Yes")
        no_cmt   = sum(1 for r in org_rows if r["_sort_key"] == NO_COMMIT)
        archived = sum(1 for r in org_rows if r["is_archived"])
        forks    = sum(1 for r in org_rows if r["is_fork"])

        for k, v in zip(
            ["scanned","b7","b30","b_mid","stale","no_cmt","archived","forks"],
            [scanned, b7, b30, b_mid, stale, no_cmt, archived, forks],
        ):
            grand[k] += v

        per_org.add_row(
            org, str(scanned),
            _c(b7,       "bold green"),
            _c(b30,      "bold yellow"),
            _c(b_mid,    "bold orange3"),
            _c(stale,    "bold red"),
            _c(no_cmt,   "dim"),
            _c(archived, "bold yellow"),
            _c(forks,    "dim cyan"),
        )

    per_org.add_row(
        "[bold white]TOTAL[/]",
        f"[bold white]{grand['scanned']}[/]",
        _c(grand["b7"],       "bold green"),
        _c(grand["b30"],      "bold yellow"),
        _c(grand["b_mid"],    "bold orange3"),
        _c(grand["stale"],    "bold red"),
        _c(grand["no_cmt"],   "dim"),
        _c(grand["archived"], "bold yellow"),
        _c(grand["forks"],    "dim cyan"),
    )
    console.print(per_org)
    console.print()

    # Stat cards
    cards = Table(box=box.SIMPLE_HEAVY, border_style="bright_blue",
                  show_header=False, expand=True, padding=(0, 2))
    for _ in range(6):
        cards.add_column("x", justify="center")

    def card(label: str, value: str, color: str) -> str:
        return f"[dim]{label}[/]\n[{color}]{value}[/{color}]"

    cards.add_row(
        card("Orgs / owners",     str(len(org_labels)),    "bold white"),
        card("Total repos",       str(grand["scanned"]),   "bold white"),
        card("≤7d active",        str(grand["b7"]),        "bold green"),
        card(f">{stale_days}d stale", str(grand["stale"]),"bold red"),
        card("Archived",          str(grand["archived"]),  "bold yellow"),
        card("API remaining",     rate_summary,            "dim cyan"),
    )
    console.print(cards)
    console.print()

    def _fmt(s: float) -> str:
        m, sec = divmod(int(s), 60)
        return f"{m}m {sec}s" if m else f"{sec:.0f}s"

    console.print(Panel(
        f"[bold green]✓  Analysis Complete[/]   [dim]·[/]   "
        f"[white]Orgs/owners:[/] [bold]{len(org_labels)}[/]   "
        f"[white]Repos:[/] [bold]{grand['scanned']}[/]   "
        f"[white]≤7d:[/] [bold green]{grand['b7']}[/]   "
        f"[white]Stale:[/] [bold red]{grand['stale']}[/]   "
        f"[white]Archived:[/] [bold yellow]{grand['archived']}[/]   "
        f"[white]No commits:[/] [dim]{grand['no_cmt']}[/]   "
        f"[white]Runtime:[/] [bold cyan]{_fmt(elapsed)}[/]",
        border_style="bright_green", box=box.DOUBLE_EDGE, expand=False,
    ))
    console.print()


# ── Export ────────────────────────────────────────────────────────────────────

# LOSS-02: fixed JSON key mapping
_JSON_KEYS = {
    "org_name":         "org_name",
    "repo_name":        "repo_name",
    "repo_url":         "repo_url",
    "is_fork":          "is_fork",
    "is_archived":      "is_archived",
    "last_commit_days": "last_commit_days",
    "last_commit_iso":  "last_commit_iso",
    "stale":            "stale",
}

COL_HEADERS = [
    "Sno", "Org / Owner", "Repo Name", "Repo URL",
    "Fork", "Archived", "Last Commit (days)", "Stale",
]

COL_WIDTHS = {
    "A": 6,   # Sno
    "B": 22,  # Org
    "C": 30,  # Repo Name
    "D": 55,  # URL
    "E": 8,   # Fork
    "F": 10,  # Archived
    "G": 20,  # Last Commit (days)
    "H": 11,  # Stale
}


def _row_to_cells(i: int, row: RepoRow) -> List:
    """Convert a RepoRow to a list of Excel cell values."""
    days = row["last_commit_days"]
    # "Last Commit (days)" column: plain integer or label string
    if days == NO_COMMIT:
        days_val = "No commits"
    elif days == 0:
        days_val = 0
    else:
        days_val = days   # plain integer — no "days ago" suffix

    return [
        i,
        row["org_name"],
        row["repo_name"],
        row["repo_url"],
        str(row["is_fork"]),
        str(row["is_archived"]),
        days_val,
        row["stale"],
    ]


def _stale_fill(stale_val: str) -> PatternFill:
    if stale_val == "Yes":        return _FILL_STALE
    if stale_val == "N/A":        return _FILL_ARCHIVED
    if stale_val == "No commits": return _FILL_WARN
    return _FILL_FRESH


def _style_ws(ws, data_rows: List[List]) -> None:
    """Apply header + data styling to a worksheet. PERF-03: single pass."""
    # Column widths
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Header
    for cell in ws[1]:
        cell.fill      = _HDR_FILL
        cell.font      = _HDR_FONT
        cell.border    = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    stale_col_idx = COL_HEADERS.index("Stale")   # 0-based for cell list

    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        stale_val = str(row_cells[stale_col_idx].value or "")
        fill      = _stale_fill(stale_val)

        for ci, cell in enumerate(row_cells):
            cell.font      = _BODY_FONT
            cell.border    = _BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            # LOSS-01: colour only the Stale column cell, not the URL cell
            if ci == stale_col_idx:
                cell.fill = fill
            else:
                cell.fill = _FILL_NONE

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def export_xlsx(rows: List[RepoRow], path: str, org_labels: List[str]) -> str:
    """
    PERF-03: single-pass openpyxl write (no pandas intermediate).
    FEAT-06: one sheet per org + "All Repos" summary sheet.
    """
    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    def _write_sheet(ws, sheet_rows: List[RepoRow]) -> None:
        ws.append(COL_HEADERS)
        for i, row in enumerate(sheet_rows, start=1):
            ws.append(_row_to_cells(i, row))
        _style_ws(ws, [])

    # All-repos sheet first
    ws_all = wb.create_sheet("All Repos")
    _write_sheet(ws_all, rows)

    # Per-org sheets
    import re
    for org in org_labels:
        org_rows = [r for r in rows if r["org_name"] == org]
        if not org_rows:
            continue
        sname = re.sub(r'[\\/*?:\[\]]', '-', org)[:31]
        ws_org = wb.create_sheet(sname)
        _write_sheet(ws_org, org_rows)

    wb.save(path)
    return os.path.abspath(path)


def export_csv(rows: List[RepoRow], path: str) -> str:
    if not rows:
        console.print("  [dim yellow]⚠  No rows to export — CSV not written.[/]")
        return path
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(COL_HEADERS)
        for i, row in enumerate(rows, start=1):
            writer.writerow(_row_to_cells(i, row))
    return os.path.abspath(path)


def export_json(rows: List[RepoRow], path: str) -> str:
    out = []
    for i, row in enumerate(rows, start=1):
        # LOSS-02: use fixed mapping, not ad-hoc _snake()
        days = row["last_commit_days"]
        out.append({
            "sno":              i,
            "org_name":         row["org_name"],
            "repo_name":        row["repo_name"],
            "repo_url":         row["repo_url"],
            "is_fork":          row["is_fork"],
            "is_archived":      row["is_archived"],
            "last_commit_days": days if days != NO_COMMIT else None,
            "last_commit_iso":  row["last_commit_iso"] or None,
            "stale":            row["stale"],
        })
    with open(path, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    return os.path.abspath(path)


# ── Progress bar ──────────────────────────────────────────────────────────────

def make_progress() -> Progress:
    return Progress(
        SpinnerColumn(spinner_name="dots", style="bright_blue"),
        TextColumn("[bold cyan]{task.description}"),
        BarColumn(bar_width=36, style="bright_blue", complete_style="bright_green"),
        TaskProgressColumn(),
        MofNCompleteColumn(),
        TimeElapsedColumn(),
        console=console,
        transient=False,
    )


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    global VERBOSE, _shutdown

    args   = parse_args()
    VERBOSE = args.verbose

    # BUG-01: compute timestamp inside main(), not at import time
    date_file = datetime.now().strftime("%Y-%m-%dT%H-%M")

    # FEAT-03: Ctrl+C → graceful shutdown
    def _handle_sigint(sig, frame):
        global _shutdown
        _shutdown = True
        with print_lock:
            console.print(
                "\n  [bold yellow]⚠  Interrupted — saving partial results...[/]"
            )
    signal.signal(signal.SIGINT, _handle_sigint)

    t_start = time.monotonic()
    print_banner(args)

    # ── Auth ──────────────────────────────────────────────────────────
    with console.status("  [dim]Resolving GitHub token...[/]", spinner="dots"):
        token = get_token()

    client = GHClient(token)

    me = client.get_one("/user")
    if not me:
        console.print("[bold red]✗  Could not fetch user info — check token.[/]")
        sys.exit(1)

    console.print(
        f"  [green]✓[/]  Logged in as [bold cyan]{me.get('login')}[/]"
        f"  [dim]({me.get('name', '')})[/]"
    )
    console.print()

    client.check_scopes()       # FEAT-07
    client.check_rate_limit()
    console.print()

    # ── Discover orgs ─────────────────────────────────────────────────
    if args.org:
        orgs = args.org
        console.print(
            f"  [green]✓[/]  Targeting [bold]{len(orgs)}[/] specified org(s): "
            f"[dim]{', '.join(orgs)}[/]"
        )
    else:
        with console.status("  [dim]Fetching organisations...[/]", spinner="dots"):
            orgs_raw, _ = client.paginate("/user/orgs")
            orgs = [o["login"] for o in orgs_raw]
        if not orgs:
            console.print(Panel(
                "[bold yellow]⚠  No organisations found.[/]",
                border_style="yellow", box=box.ROUNDED, expand=False))
            if not args.include_personal:
                console.print("  [dim]Tip: try [bold]--include-personal[/][/]")
                sys.exit(0)
        else:
            console.print(
                f"  [green]✓[/]  Found [bold]{len(orgs)}[/] organisation(s):  "
                f"[dim]{', '.join(orgs)}[/]"
            )
    console.print()

    # ── Collect repos — PERF-01: orgs fetched concurrently ─────────
    repo_queue: List[Tuple[str, dict]] = []
    seen_full_names: Set[str] = set()   # BUG-05: dedup set

    with make_progress() as progress:
        task = progress.add_task("[cyan]Listing repositories...", total=len(orgs))
        with ThreadPoolExecutor(max_workers=min(len(orgs), 5)) as pool:
            fut_map: Dict[Future, str] = {
                pool.submit(
                    fetch_org_repos,
                    client, org,
                    args.exclude_forks,
                    args.exclude_archived,
                ): org
                for org in orgs
            }
            for fut in as_completed(fut_map):
                org = fut_map[fut]
                try:
                    org_repos = fut.result()
                    for r in org_repos:
                        fn = r.get("full_name", "")
                        if fn not in seen_full_names:
                            seen_full_names.add(fn)
                            repo_queue.append((org, r))
                except Exception as exc:
                    with print_lock:
                        console.print(f"  [bold red]✗[/]  Listing {org}: {exc}")
                progress.advance(task)

    # Personal repos (BUG-05: dedup against org repos)
    if args.include_personal:
        personal_login = me.get("login", "")
        with console.status(
            f"  [dim]Listing personal repos for [bold]{personal_login}[/]...[/]",
            spinner="dots",
        ):
            personal_repos, _ = client.paginate(
                "/user/repos", params={"type": "owner", "affiliation": "owner"}
            )
        for r in personal_repos:
            if args.exclude_forks and r.get("fork"):
                continue
            if args.exclude_archived and r.get("archived"):
                continue
            fn = r.get("full_name", "")
            if fn not in seen_full_names:   # BUG-05: actual dedup check
                seen_full_names.add(fn)
                repo_queue.append((personal_login, r))
        if personal_repos:
            console.print(
                f"  [green]✓[/]  Personal repos queued: "
                f"[bold]{sum(1 for o,_ in repo_queue if o == personal_login)}[/]"
            )

    if not repo_queue:
        console.print(Panel(
            "[bold yellow]⚠  No repositories found to scan.[/]",
            border_style="yellow", box=box.ROUNDED, expand=False))
        sys.exit(0)

    console.print(
        f"  [green]✓[/]  [bold]{len(repo_queue)}[/] repo(s) queued"
        + (" (forks excluded)"    if args.exclude_forks    else "")
        + (" (archived excluded)" if args.exclude_archived else "")
    )
    console.print()

    # ── Concurrent commit fetch ────────────────────────────────────────
    all_rows: List[RepoRow] = []

    with make_progress() as progress:
        task = progress.add_task("[cyan]Fetching commit dates...", total=len(repo_queue))
        with ThreadPoolExecutor(max_workers=args.workers) as pool:
            future_map: Dict[Future, Tuple[str, str]] = {
                pool.submit(fetch_repo_row, client, org, repo, args.stale_days, args.since): (
                    org, repo.get("name", "")
                )
                for org, repo in repo_queue
            }
            for fut in as_completed(future_map):
                if _shutdown:
                    pool.shutdown(wait=False, cancel_futures=True)
                    break
                org_label, repo_name = future_map[fut]
                progress.update(
                    task,
                    description=f"[cyan]{org_label}[/] — [white]{repo_name[:28]}[/]",
                )
                try:
                    all_rows.append(fut.result())
                except InterruptedError:
                    pass
                except Exception as exc:
                    with print_lock:
                        console.print(
                            f"  [bold red]✗[/]  Error: "
                            f"[bold]{org_label}/{repo_name}[/]: {exc}"
                        )
                progress.advance(task)

    if not all_rows:
        console.print("[bold yellow]⚠  No results collected.[/]")
        sys.exit(0)

    # Sort
    all_rows = sort_rows(all_rows, args.sort_by)

    # Collect unique org labels preserving order
    seen_orgs: Set[str] = set()
    org_labels: List[str] = []
    for r in all_rows:
        if r["org_name"] not in seen_orgs:
            seen_orgs.add(r["org_name"])
            org_labels.append(r["org_name"])

    # ── Terminal output ───────────────────────────────────────────────
    print_results_table(all_rows, args.filter_mode, args.stale_days)

    # ── Export ────────────────────────────────────────────────────────
    if not args.no_export:
        ext         = args.output_format
        default_name = f"commit_analyser_{date_file}.{ext}"
        out_path    = args.output or default_name

        with console.status(
            f"[bright_blue]Writing {ext.upper()} report...[/]", spinner="dots"
        ):
            if ext == "xlsx":
                saved = export_xlsx(all_rows, out_path, org_labels)
            elif ext == "csv":
                saved = export_csv(all_rows, out_path)
            else:
                saved = export_json(all_rows, out_path)

        console.print(f"  [green]✓[/]  {ext.upper()} saved → [dim]{saved}[/]")
        console.print()

    # ── Summary ───────────────────────────────────────────────────────
    elapsed = time.monotonic() - t_start
    print_summary(
        all_rows, org_labels, elapsed,
        args.stale_days,
        client.rate_summary(),   # QUAL-04: pass string not GHClient
    )

    if _shutdown:
        console.print("  [bold yellow]⚠  Scan was interrupted — output may be partial.[/]")
        console.print()


if __name__ == "__main__":
    main()
